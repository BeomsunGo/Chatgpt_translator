import openai
import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from ast import literal_eval
import sys
import requests
import time
import re
from io import BytesIO
import base64

class limit_error(Exception):
    pass

def import_excel(file_path):

    wb = Workbook()
    if file_path.name[-4:] == "xlsm":
        wb = load_workbook(file_path,data_only=True,keep_vba=True)
    elif file_path.name[-4:] == "xlsx":
        wb = load_workbook(file_path,data_only=True)
    else :
        st.write("파일형식오류")
        sys.exit()

    return wb

def make_dict(ws_list,org_lang):

    trans_dict = {}
    for order, wsname in enumerate(ws_list):
        ws = wb[wsname]
        max_row = ws.max_row
        max_col = ws.max_column
        st.write("시트명 : "+wsname+", 최대 행 수 : "+str(max_row)+", 최대 열 수 : "+str(max_col))
        for row in range(1,max_row+1):
            for col in range(1, max_col+1):
                target = ws.cell(row, col).value
                if target == None:
                    continue
                if is_not_org_sentence(str(target),org_lang):
                    continue

                key = str(order)+"-"+str(row)+"-"+str(col)
                val = target

                trans_dict[key] = val

    return trans_dict

def make_trans_DB(string,df):

    trans_DB = {}
    if df.empty:
        pass
    else:
        for word in df["번역전"]:
            if word in string:
                index = df.index[df["번역전"]==word][0]
                trans_DB[word] = df.loc[index,"번역후"]

    return trans_DB

def slice_dict(dict, max_length,df):

    result = []
    result_DB = []
    current_dict = {}
    current_length = 0
    current_trans_DB = {}
    tot_cnt = 1

    for key, value in dict.items():
        key_length = len(str(key))
        value_length = len(str(value))
        trans_DB_length = len(str(current_trans_DB))

        if df.empty:
            trans_DB_length = 0
        else:
            trans_DB_length = len(str(current_trans_DB))

        if current_length +key_length + value_length + trans_DB_length > max_length:
            result.append(current_dict)
            result_DB.append(current_trans_DB)
            current_dict = {}
            current_trans_DB = {}
            current_length = 0
            tot_cnt += 1
        current_dict[key] = value
        current_length += key_length
        current_length += value_length
        current_trans_DB.update(make_trans_DB(value,df))
    result.append(current_dict)
    result_DB.append(current_trans_DB)
    return result, result_DB, tot_cnt


def do_translate(messages):
    completions = openai.ChatCompletion.create(
      # model="gpt-4",
        model="gpt-3.5-turbo-16k",
        messages=messages,
        timeout=60
    )
#     st.write("used token :"+str(completions.usage['total_tokens']))
    if completions.usage['total_tokens']>16384:
        raise limit_error

    return completions

def is_not_org_sentence(text,org_lang):
    # Only contains english alphabet and several marks include 'space'.
    check_num = 0
    if org_lang == "English":
        pattern = re.compile(r'[a-zA-Z]')
    elif org_lang == "Korean":
        pattern = re.compile(r'[ㄱ-ㅎ가-힣]')
    else:
        pattern = re.compile(r'.*')
    for char in text:
        if pattern.match(char):
            check_num+=1
            break
    if check_num == 0:
        return True
    else:
        return False



############### main ################

openai.api_key = st.secrets["OPENAI_KEY"]


st.set_page_config(layout="wide")

st.title('Assurance DA')
st.header('File Translator')
st.write('Developed by Assurance DA (beomsun.go@pwc.com)')



col1, col2 = st.columns([1,2])

with col1:
    org_lang = st.radio("Input 언어를 선택하세요", ["Korean", "English", "Chinese", "Japanese"], horizontal=True)
    tobe_lang = st.radio("Output 언어를 선택하세요", ["Korean", "English", "Chinese", "Japanese"], horizontal=True)


df_empty = pd.DataFrame(columns = ['번역전','번역후'])
df = pd.DataFrame()

with col1:
    DB_type = st.radio(
        "지정된 번역을 사용할 단어를 입력하세요",
        ("엑셀파일","직접입력"), horizontal=True
    )

    if DB_type == "엑셀파일":
        file_DB = st.file_uploader(
            "파일을 선택하세요(xlsx만 가능)",
            type=['xlsx']
        )
        df_count_bef = 0
        df_count_aft = 0
        if file_DB is not None:
            df = pd.read_excel(file_DB,engine="openpyxl")
        else:
            df = pd.DataFrame(columns = ['번역전','번역후'])

    elif DB_type == "직접입력":
        df = st.experimental_data_editor(df_empty, use_container_width = True,num_rows="dynamic")
        df_count = df.shape[0]
        df_count_bef = df["번역전"].count()
        df_count_aft = df["번역후"].count()
        st.write(f' - 번역 전 단어 : {df_count_bef}개')
        st.write(f' - 번역 후 단어 : {df_count_aft}개')
    else:
        df = pd.DataFrame(columns = ['번역전','번역후'])
with col2:
    file = st.file_uploader(
        "파일을 선택하세요(xlsx, xlsm만 가능)",
        type=['xlsx', 'xlsm']
    )


    if (file is not None and df_count_bef == df_count_aft) and st.button("번역 시작"):

        st.write(file.name)
        # file_path = r"C:\Users\bgo006\Desktop\CorDA\project\chatgpt\translator\sample_eng_2.xlsm"
        # lang = "English"
        # st.write(lang)
        #################### 엑셀 불러온 후 모든 글자 긁어오기 #####################
        wb = import_excel(file_path=file)
        st.write("The Excel file has been uploaded.")
        ws_list = wb.sheetnames


        #################### dictionary 생성 ###################
        trans_dict = make_dict(ws_list = ws_list,org_lang=org_lang)
        st.write("Excel data has been loaded.")

        ###################### text limit 설정(환경이나 상황 등에 맞춰 적절한 값으로 바꿔줘야함) ###################
        if org_lang == "Korean" and tobe_lang == "English":
            text_limit = 4100
        elif org_lang == "English" and tobe_lang == "Korean":
            text_limit = 4100
        else:
            text_limit = 3000

        sliced_dicts, sliced_DB_dicts, tot_cnt = slice_dict(trans_dict,text_limit,df) # 한자는 1,300자로 하는게 안전한듯 # 영어는 2500자?
        st.write("Input dictionaries have been created.")


        answer_dicts = {}
        st.write("번역시작")
        for trytime, sliced_dict in enumerate(sliced_dicts):
            messages = []
            sliced_trans_DB = sliced_DB_dicts[trytime]
            st.write(f"Input : {trytime+1}/{tot_cnt}")
#             st.write(f"Input 길이 : {len(str(sliced_dict))}")
#             st.write(sliced_trans_DB)
#             st.write(str(sliced_dict))
            messages.append({"role": "system", "content": 'You are a translate program. Dictionary is one of the type of variables in python that contains keys and values. The beginning and end of a dictionary are represented by \'{\' and \'}\', respectively, and the key and value are connected by \':\'. Each key-value pair separated by \', \' with no other spaces or line break. Also, There are any space or line break between \"{\" and first key-value pair, \"}\" and last key-value pair respectively'})
            # messages.append({"role": "system", "content": 'Please translate sentenses and words from English to Korean. What you should translate are values in below dictionary and output type is also dictionary which has same keys with input dictionary'})
            messages.append({"role": "system", "content": f'Translate all the {org_lang} words and sentences in the dictionary below target dictionary into {tobe_lang}. What you should translate are all the sentenses and words. Output type is also dictionary which has same keys with input dictionary.'})
            messages.append({"role": "system", "content": f'{str(sliced_dict)}'})
            if not df.empty:
                messages.append({"role": "system", "content": f'Use the following dictionary when translating -> {str(sliced_trans_DB)}.'})
            # messages.append({"role": "system", "content": f'If there is \' or \" with in the middle of the translated sentence, replace then with \\\' , \\\".'})
            messages.append({"role": "system", "content": f'Output should be only an Dictionary without any comments.'})


            try:

                try:

#                     st.write("try : 1")

                    completions = do_translate(messages=messages)

                    answer = completions.choices[0]['message']['content']


#                     st.write(answer)

                    answer = answer.replace('\'s','\\\'s')

                    answer = answer.replace('\\\\\'s','\\\'s')

                    answer_dict = literal_eval(answer)

                    answer_dicts.update(answer_dict)


#                     st.write("try : 1 - finish")

                except requests.exceptions.Timeout:

                    time.sleep(2)

#                     st.write("try : 2 - timeout")

                    completions = do_translate(messages=messages)

                    answer = completions.choices[0]['message']['content']


#                     st.write(answer)

                    answer = answer.replace('\'s','\\\'s')

                    answer = answer.replace('\\\\\'s','\\\'s')

                    answer_dict = literal_eval(answer)

                    answer_dicts.update(answer_dict)

#                     st.write("try : 2 - Finish")

                except SyntaxError:

                    try:

                        time.sleep(2)

#                         st.write("try : 2 - syntax")

                        completions = do_translate(messages=messages)

                        answer = completions.choices[0]['message']['content']


#                         st.write(answer)

                        answer = answer.replace('\'s','\\\'s')

                        answer = answer.replace('\\\\\'s','\\\'s')

                        answer_dict = literal_eval(answer)

                        answer_dicts.update(answer_dict)

#                         st.write("try : 2 - Finish")


                    except SyntaxError:

                        time.sleep(2)

#                         st.write("try : 3 - syntax")

                        completions = do_translate(messages=messages)

                        answer = completions.choices[0]['message']['content']


#                         st.write(answer)

                        answer = answer.replace('\'s','\\\'s')

                        answer = answer.replace('\\\\\'s','\\\'s')

                        answer_dict = literal_eval(answer)

                        answer_dicts.update(answer_dict)

#                         st.write("try : 3 - Finish")


                    except limit_error:

                        st.write("해당 셀에 너무 긴 문장이 들어 있습니다.")
                        sent_vals = list(sliced_dict.values())

                        sent_idxs = list(sliced_dict.keys())

                        for sent_idx, sent_val in enumerate(sent_vals):
                            ans_sent = []
                            answer_dict = {}
                            sent_splits = sent_val.split(". ")

                            for sent_split in sent_splits:
                                messages = []
                                messages.append({"role": "system", "content": 'You are a translate program.'})
                                messages.append({"role": "system", "content": f'Please translate this sentencs. \'{sent_split}\''})
                                completions = do_translate(messages=messages)
                                answer = completions.choices[0]['message']['content']

                                answer = answer.replace('\'s','\\\'s')
                                answer = answer.replace('\\\\\'s','\\\'s')
                                ans_sent.append(answer)

                            ans_sent_tot = ".".join(ans_sent)

#                             st.write(ans_sent_tot)

                            answer_dict[sent_idx] = ans_sent_tot

                            answer_dicts.update(answer_dict)




                except limit_error:

                    st.write("해당 셀에 너무 긴 문장이 들어 있습니다.")
                    sent_vals = list(sliced_dict.values())

                    sent_idxs = list(sliced_dict.keys())

                    for sent_idx, sent_val in enumerate(sent_vals):
                        ans_sent = []
                        answer_dict = {}
                        sent_splits = sent_val.split(". ")

                        for sent_split in sent_splits:
                            messages = []
                            messages.append({"role": "system", "content": 'You are a translate program.'})
                            messages.append({"role": "system", "content": f'Please translate this sentencs. \'{sent_split}\''})
                            completions = do_translate(messages=messages)
                            answer = completions.choices[0]['message']['content']

                            answer = answer.replace('\'s','\\\'s')
                            answer = answer.replace('\\\\\'s','\\\'s')
                            ans_sent.append(answer)

                        ans_sent_tot = ".".join(ans_sent)

#                         st.write(ans_sent_tot)

                        answer_dict[sent_idx] = ans_sent_tot

                        answer_dicts.update(answer_dict)

            except :

                st.write("오류로 인해 해당부분이 번역되지 않았습니다.")

        for key_answer in answer_dicts:
            try:
                val_answer = answer_dicts[key_answer]
                key_answer_list = key_answer.split("-")
                wsname_answer = ws_list[int(key_answer_list[0])]
                row_answer = int(key_answer_list[1])
                col_answer = int(key_answer_list[2])
                wb[wsname_answer].cell(row_answer,col_answer).value = val_answer
           
            except:
                continue



        st.write("번역완료")

        #### output 생성 ####
        output = BytesIO()
        output_file_name = f"{'.'.join(file.name.split('.')[0:-1])}_{tobe_lang}.{file.name.split('.')[-1]}"
        wb.save(output)
        output_file = output.getvalue()
        b64 = base64.b64encode(output_file)
        download_link = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64.decode()}" download=\'{output_file_name}\'>Download Excel File</a>'
        st.write("파일 생성 완료")
        st.markdown(download_link, unsafe_allow_html=True)    

        time.sleep(100)
